import io
import pathlib
import shutil
import zipfile
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter


def _update_checkboxes(out_path: str, sheet_position_updates: dict[str, dict[tuple, bool]]) -> None:
    """
    Modify <x:Checked> in VML XML for checkboxes identified by (row, col).

    sheet_position_updates: {sheet_name: {(row, col): True/False}}
    """
    from lxml import etree
    from excel_read_helpers import _find_sheet_path, _find_vml_files_for_sheet

    ns_v = "urn:schemas-microsoft-com:vml"
    ns_x = "urn:schemas-microsoft-com:office:excel"

    # Read all zip contents up front
    with zipfile.ZipFile(out_path, "r") as z:
        all_names = z.namelist()
        file_contents = {name: z.read(name) for name in all_names}

    modified: dict[str, bytes] = {}

    for sheet_name, position_updates in sheet_position_updates.items():
        if not position_updates:
            continue

        with zipfile.ZipFile(out_path, "r") as z:
            sheet_path = _find_sheet_path(z, sheet_name)
            if not sheet_path:
                continue
            vml_files = _find_vml_files_for_sheet(z, sheet_path)

        for vml_file in vml_files:
            if vml_file not in file_contents:
                continue

            root = etree.fromstring(file_contents[vml_file], etree.XMLParser(recover=True))
            changed = False

            for shape in root.findall(f".//{{{ns_v}}}shape"):
                cd = shape.find(f".//{{{ns_x}}}ClientData")
                if cd is None or cd.get("ObjectType") != "Checkbox":
                    continue

                anchor_el = cd.find(f"{{{ns_x}}}Anchor")
                if anchor_el is None or not anchor_el.text:
                    continue

                parts = [p.strip() for p in anchor_el.text.split(",")]
                row, col = 0, 0
                if len(parts) >= 8:
                    col = int(parts[0]) + 1
                    row = ((int(parts[2]) + int(parts[6])) // 2) + 1
                elif len(parts) >= 4:
                    col = int(parts[0]) + 1
                    row = int(parts[2]) + 1

                if (row, col) not in position_updates:
                    continue

                should_check = position_updates[(row, col)]
                checked_el = cd.find(f"{{{ns_x}}}Checked")

                if should_check:
                    if checked_el is None:
                        checked_el = etree.SubElement(cd, f"{{{ns_x}}}Checked")
                    checked_el.text = "1"
                else:
                    if checked_el is not None:
                        cd.remove(checked_el)

                changed = True

            if changed:
                modified[vml_file] = etree.tostring(
                    root, xml_declaration=True, encoding="UTF-8", standalone=True
                )

    if not modified:
        return

    # Rewrite the zip with the modified VML files
    buf = io.BytesIO()
    with zipfile.ZipFile(out_path, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                zout.writestr(name, modified.get(name, zin.read(name)))

    with open(out_path, "wb") as f:
        f.write(buf.getvalue())


def _update_ctrlprops(out_path: str, sheet_position_updates: dict[str, dict[tuple, bool]]) -> None:
    """
    Modify checked="Checked" in ctrlProp XML files for modern xlsx checkboxes.

    sheet_position_updates: {sheet_name: {(row, col): True/False}}
    """
    import posixpath
    from lxml import etree
    from excel_read_helpers import _find_sheet_path, _find_vml_files_for_sheet

    ns_v = "urn:schemas-microsoft-com:vml"
    ns_x = "urn:schemas-microsoft-com:office:excel"
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    with zipfile.ZipFile(out_path, "r") as z:
        all_names = z.namelist()
        file_contents = {name: z.read(name) for name in all_names}

    modified: dict[str, bytes] = {}

    for sheet_name, position_updates in sheet_position_updates.items():
        if not position_updates:
            continue

        with zipfile.ZipFile(out_path, "r") as z:
            sheet_path = _find_sheet_path(z, sheet_name)
            if not sheet_path:
                continue

            parts = sheet_path.rsplit("/", 1)

            # rId -> ctrlProp file path from sheet .rels
            rels_path = f"{parts[0]}/_rels/{parts[1]}.rels"
            rid_to_ctrlprop: dict[str, str] = {}
            if rels_path in z.namelist():
                rels_root = etree.fromstring(z.read(rels_path))
                for rel in rels_root.findall(f"{{{ns_pkg}}}Relationship"):
                    target = rel.get("Target", "")
                    if "ctrlProp" in target:
                        resolved = posixpath.normpath(f"{parts[0]}/{target}")
                        rid_to_ctrlprop[rel.get("Id")] = resolved

            # shapeId (numeric string) -> rId from sheet XML
            sheet_root = etree.fromstring(z.read(sheet_path))
            shapeid_to_rid: dict[str, str] = {}
            for el in sheet_root.iter():
                if el.tag.endswith("}control") or el.tag == "control":
                    sid = el.get("shapeId")
                    rid = el.get(f"{{{ns_r}}}id")
                    if sid and rid:
                        shapeid_to_rid[sid] = rid

            vml_files = _find_vml_files_for_sheet(z, sheet_path)

        # VML anchor -> (row, col) -> ctrlProp path + desired state
        pos_to_ctrlprop: dict[tuple, tuple[str, bool]] = {}
        for vml_file in vml_files:
            if vml_file not in file_contents:
                continue
            vml_root = etree.fromstring(file_contents[vml_file], etree.XMLParser(recover=True))
            for shape in vml_root.findall(f".//{{{ns_v}}}shape"):
                shape_id_attr = shape.get("id", "")
                if "_x0000_s" not in shape_id_attr:
                    continue
                shape_num = shape_id_attr.split("_x0000_s")[-1]

                cd = shape.find(f".//{{{ns_x}}}ClientData")
                if cd is None or cd.get("ObjectType") != "Checkbox":
                    continue

                anchor_el = cd.find(f"{{{ns_x}}}Anchor")
                if anchor_el is None or not anchor_el.text:
                    continue

                anchor_parts = [p.strip() for p in anchor_el.text.split(",")]
                row, col = 0, 0
                if len(anchor_parts) >= 8:
                    col = int(anchor_parts[0]) + 1
                    row = ((int(anchor_parts[2]) + int(anchor_parts[6])) // 2) + 1
                elif len(anchor_parts) >= 4:
                    col = int(anchor_parts[0]) + 1
                    row = int(anchor_parts[2]) + 1

                if (row, col) not in position_updates:
                    continue

                rid = shapeid_to_rid.get(shape_num)
                if rid:
                    ctrlprop_path = rid_to_ctrlprop.get(rid)
                    if ctrlprop_path:
                        pos_to_ctrlprop[(row, col)] = (ctrlprop_path, position_updates[(row, col)])

        for (row, col), (ctrlprop_path, should_check) in pos_to_ctrlprop.items():
            if ctrlprop_path not in file_contents:
                continue
            ctrlprop_root = etree.fromstring(file_contents[ctrlprop_path])
            if should_check:
                ctrlprop_root.set("checked", "Checked")
            else:
                ctrlprop_root.attrib.pop("checked", None)
            modified[ctrlprop_path] = etree.tostring(
                ctrlprop_root, xml_declaration=True, encoding="UTF-8", standalone=True
            )

    if not modified:
        return

    buf = io.BytesIO()
    with zipfile.ZipFile(out_path, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                zout.writestr(name, modified.get(name, zin.read(name)))

    with open(out_path, "wb") as f:
        f.write(buf.getvalue())


def _restore_after_openpyxl_save(
    out_path: str,
    pre_save: dict[str, bytes],
) -> None:
    """
    After openpyxl saves (and strips VML, ctrlProps, sheet rels, controls elements),
    restore everything it removed:
      1. Files present in pre_save but missing in the saved zip → add back verbatim.
      2. For EVERY sheet XML that had <controls>, <legacyDrawing>, or <mc:AlternateContent>
         stripped by openpyxl, patch those elements back from the pre-save version.
         This must cover all sheets, not just the one being written — openpyxl rewrites
         the entire workbook on every save, stripping these elements from every sheet.
    """
    from lxml import etree

    ns_ss = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_mc = "http://schemas.openxmlformats.org/markup-compatibility/2006"
    tags_to_restore = {
        f"{{{ns_ss}}}controls",
        f"{{{ns_ss}}}legacyDrawing",
        f"{{{ns_ss}}}legacyDrawingHF",
        f"{{{ns_mc}}}AlternateContent",  # wraps <controls> in modern xlsx
    }

    with zipfile.ZipFile(out_path, "r") as z:
        post_names = set(z.namelist())
        post_contents = {name: z.read(name) for name in z.namelist()}

    missing = {name: data for name, data in pre_save.items() if name not in post_names}
    patched: dict[str, bytes] = {}

    # Scan every sheet XML in pre_save; restore any VML-related elements openpyxl stripped.
    for name, pre_data in pre_save.items():
        if "worksheets/sheet" not in name or not name.endswith(".xml"):
            continue
        post_data = post_contents.get(name)
        if post_data is None or post_data == pre_data:
            continue  # missing handled below, or genuinely unchanged

        pre_root = etree.fromstring(pre_data)
        elements_to_restore = [el for el in pre_root if el.tag in tags_to_restore]
        if not elements_to_restore:
            continue  # this sheet had no VML elements — nothing to restore

        post_root = etree.fromstring(post_data)
        for pre_el in elements_to_restore:
            for existing in post_root.findall(pre_el.tag):
                post_root.remove(existing)
            post_root.append(pre_el)

        # Merge namespace maps so prefixes like r:, mc:, xdr: are declared at the root.
        merged_nsmap = {**pre_root.nsmap, **post_root.nsmap}
        if merged_nsmap != post_root.nsmap:
            new_root = etree.Element(post_root.tag, nsmap=merged_nsmap)
            new_root.attrib.update(post_root.attrib)
            new_root.text = post_root.text
            new_root.tail = post_root.tail
            for child in list(post_root):
                new_root.append(child)
            post_root = new_root

        patched[name] = etree.tostring(
            post_root, xml_declaration=True, encoding="UTF-8", standalone=True
        )

    # Restore files that openpyxl modified but we didn't intentionally change:
    #   - [Content_Types].xml: openpyxl regenerates it, stripping entries for VML/ctrlProps/etc.
    #   - drawing*.xml: openpyxl strips checkbox shapes while keeping the file
    for name, pre_data in pre_save.items():
        post_data = post_contents.get(name)
        if post_data is None or post_data == pre_data:
            continue  # missing (handled separately) or unchanged
        if (name == "[Content_Types].xml"
                or ("drawings/drawing" in name and "vml" not in name.lower())):
            patched[name] = pre_data

    if not missing and not patched:
        return

    buf = io.BytesIO()
    with zipfile.ZipFile(out_path, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                # Drop drawing rels files that openpyxl added — they weren't in the
                # original and cause Excel to flag drawing1.xml as broken.
                if name.endswith(".rels") and "drawings/_rels" in name and name not in pre_save:
                    continue
                zout.writestr(name, patched.get(name, zin.read(name)))
            for name, data in missing.items():
                zout.writestr(name, data)

    with open(out_path, "wb") as f:
        f.write(buf.getvalue())


def write_cells(
    file_path: str,
    cells: dict[str, list[dict]],
    value: str | None = None,
    out_path: str | None = None,
) -> str:
    """
    Write values into cells, handling both regular input cells and VML checkboxes.

    Parameters
    ----------
    file_path : str
        Source workbook path.
    cells : dict[sheet_name -> list[cell_dict]]
        Same structure returned by extract_cells — each dict must have
        "sheet", "cell", and "type" keys. When *value* is None, each dict
        must also have a "value" key; cells whose "value" is None are skipped.
    value : str | None
        Fixed value to write into every cell. If None, uses item["value"]
        from each cell dict instead.
    out_path : str | None
        Destination path. If None, overwrites file_path in-place.

    Returns
    -------
    str
        Path of the written file.
    """
    out_path = out_path or file_path

    if out_path != file_path:
        shutil.copy2(file_path, out_path)

    # Separate input cells from checkbox cells
    input_writes: dict[str, list[tuple[str, str]]] = {}   # sheet -> [(coord, val)]
    checkbox_writes: dict[str, list[tuple[str, str]]] = {}  # sheet -> [(coord, val)]

    for sheet_name, cell_list in cells.items():
        for item in cell_list:
            cell_value = value if value is not None else item.get("value")
            if cell_value is None:
                continue
            coord = item["cell"]
            if item.get("type") == "checkbox":
                checkbox_writes.setdefault(sheet_name, []).append((coord, cell_value))
            else:
                input_writes.setdefault(sheet_name, []).append((coord, cell_value))

    # Snapshot zip before openpyxl touches it (needed to restore stripped files)
    pre_save: dict[str, bytes] = {}
    with zipfile.ZipFile(out_path, "r") as z:
        pre_save = {name: z.read(name) for name in z.namelist()}

    # --- Write regular input cells via openpyxl ---
    wb = openpyxl.load_workbook(out_path)
    written = 0
    for sheet_name, writes in input_writes.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [write]  WARNING: sheet '{sheet_name}' not found — skipped")
            continue
        ws = wb[sheet_name]
        for coord, val in writes:
            ws[coord] = val
            written += 1
    if written > 0:
        wb.save(out_path)
        # openpyxl strips VML, ctrlProps, sheet rels, and controls/legacyDrawing from
        # sheet XMLs — restore them from the pre-save snapshot
        _restore_after_openpyxl_save(out_path, pre_save)
    wb.close()

    # --- Write checkboxes ---
    if checkbox_writes:
        from excel_read_helpers import extract_checkboxes

        # Build (row, col) -> should_check map for VML update
        # Also collect linked cells to update via openpyxl
        sheet_position_updates: dict[str, dict[tuple, bool]] = {}
        linked_cell_writes: dict[str, list[tuple[str, bool]]] = {}  # sheet -> [(coord, bool)]

        for sheet_name, writes in checkbox_writes.items():
            # Re-extract checkboxes to get linked cells and verify positions
            cbs = extract_checkboxes(file_path, sheet_name)
            linked_map: dict[str, str] = {}  # coord -> linked_cell
            for cb in cbs:
                if cb["linked_cell"]:
                    coord = f"{get_column_letter(cb['col'])}{cb['row']}"
                    linked_map[coord] = cb["linked_cell"]

            for coord, val in writes:
                should_check = val in (1, "1", True)
                col_str, row = coordinate_from_string(coord)
                col = column_index_from_string(col_str)
                sheet_position_updates.setdefault(sheet_name, {})[(row, col)] = should_check

                linked = linked_map.get(coord)
                if linked:
                    if "!" in linked:
                        linked = linked.split("!")[-1]
                    linked_cell_writes.setdefault(sheet_name, []).append((linked, should_check))

            written += len(writes)

        # Update VML XML and ctrlProp XML (modern xlsx checkboxes)
        _update_checkboxes(out_path, sheet_position_updates)
        _update_ctrlprops(out_path, sheet_position_updates)

        # Update linked cells via openpyxl.
        # Snapshot first — openpyxl save will strip VML/ctrlProps that
        # _update_checkboxes/_update_ctrlprops just patched in.
        if linked_cell_writes:
            with zipfile.ZipFile(out_path, "r") as z:
                pre_lc: dict[str, bytes] = {name: z.read(name) for name in z.namelist()}

            wb = openpyxl.load_workbook(out_path)
            for sheet_name, lc_writes in linked_cell_writes.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                for coord, checked in lc_writes:
                    ws[coord] = checked
            wb.save(out_path)
            wb.close()

            _restore_after_openpyxl_save(out_path, pre_lc)

    print(f"  [write]  wrote {written} cells → {pathlib.Path(out_path).name}")
    return out_path
