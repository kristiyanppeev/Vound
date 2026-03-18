import openpyxl


def extract_checkboxes(file_path, sheet_name):
    """Returns list of dicts: {label, checked, row, col, linked_cell} for a specific sheet."""
    import zipfile
    from lxml import etree

    checkboxes = []

    with zipfile.ZipFile(file_path, 'r') as z:
        # Step 1: Find the target sheet's path (e.g. xl/worksheets/sheet1.xml)
        target_sheet_path = _find_sheet_path(z, sheet_name)
        if not target_sheet_path:
            return checkboxes

        # Step 2: Find VML files linked to THIS sheet only
        vml_files = _find_vml_files_for_sheet(z, target_sheet_path)

        ns_v = 'urn:schemas-microsoft-com:vml'
        ns_x = 'urn:schemas-microsoft-com:office:excel'

        for vml_file in vml_files:
            try:
                content = z.read(vml_file)
                root = etree.fromstring(content, etree.XMLParser(recover=True))

                for shape in root.findall(f'.//{{{ns_v}}}shape'):
                    cd = shape.find(f'.//{{{ns_x}}}ClientData')
                    if cd is None or cd.get('ObjectType') != 'Checkbox':
                        continue

                    # Checked?
                    checked_el = cd.find(f'{{{ns_x}}}Checked')
                    is_checked = checked_el is not None and (
                        checked_el.text or '').strip() == '1'

                    # Label
                    textbox = shape.find(f'.//{{{ns_v}}}textbox')
                    label = ""
                    if textbox is not None:
                        label = etree.tostring(
                            textbox, method='text', encoding='unicode').strip()

                    # Position from anchor
                    anchor_el = cd.find(f'{{{ns_x}}}Anchor')
                    row, col = 0, 0
                    if anchor_el is not None and anchor_el.text:
                        parts = [p.strip() for p in anchor_el.text.split(',')]
                        if len(parts) >= 8:
                            col = int(parts[0]) + 1
                            row_start = int(parts[2])
                            row_end = int(parts[6])
                            row = ((row_start + row_end) // 2) + 1
                        elif len(parts) >= 4:
                            col = int(parts[0]) + 1
                            row = int(parts[2]) + 1

                    # Linked cell
                    fmla_el = cd.find(f'{{{ns_x}}}FmlaLink')
                    linked_cell = (fmla_el.text or '').strip(
                    ) if fmla_el is not None else None

                    checkboxes.append({
                        'label': label,
                        'checked': is_checked,
                        'row': row,
                        'col': col,
                        'linked_cell': linked_cell,
                    })
            except Exception as e:
                print(f"Warning: Could not parse {vml_file}: {e}")

    return checkboxes


def _find_sheet_path(z, sheet_name=None):
    """Find xl/worksheets/sheetN.xml for the target sheet."""
    import zipfile
    from lxml import etree

    try:
        wb_xml = z.read('xl/workbook.xml')
        root = etree.fromstring(wb_xml)
        ns_s = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

        sheets = root.findall(f'.//{{{ns_s}}}sheet')

        # Map rId -> target from workbook rels
        rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_root = etree.fromstring(rels_xml)
        rid_map = {}
        for rel in rels_root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
            rid_map[rel.get('Id')] = rel.get('Target')

        # Find matching sheet
        target_sheet = None
        for sheet_el in sheets:
            name = sheet_el.get('name')
            rid = sheet_el.get(f'{{{ns_r}}}id')
            if sheet_name and name == sheet_name:
                target_sheet = rid
                break

        # Fallback to first sheet if no match
        if target_sheet is None and sheets:
            target_sheet = sheets[0].get(f'{{{ns_r}}}id')

        if target_sheet and target_sheet in rid_map:
            target = rid_map[target_sheet]
            if not target.startswith('xl/'):
                target = f'xl/{target}'
            return target

    except Exception as e:
        print(f"Warning: Could not determine sheet path: {e}")

    return None


def _find_vml_files_for_sheet(z, sheet_path):
    """Find VML drawing files referenced by this sheet's .rels file."""
    import posixpath
    from lxml import etree

    vml_files = []
    parts = sheet_path.rsplit('/', 1)
    if len(parts) != 2:
        return vml_files

    rels_path = f"{parts[0]}/_rels/{parts[1]}.rels"
    if rels_path not in z.namelist():
        return vml_files

    try:
        rels_xml = z.read(rels_path)
        rels_root = etree.fromstring(rels_xml)
        ns_pkg = 'http://schemas.openxmlformats.org/package/2006/relationships'

        for rel in rels_root.findall(f'{{{ns_pkg}}}Relationship'):
            target = rel.get('Target', '')
            if 'vmlDrawing' in target:
                resolved = posixpath.normpath(f"{parts[0]}/{target}")
                if resolved in z.namelist():
                    vml_files.append(resolved)
    except Exception as e:
        print(f"Warning: Could not parse {rels_path}: {e}")

    return vml_files


def list_tables(file_path):
    """Prints all Excel tables in each sheet of the workbook."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tables = ws.tables
        if tables:
            print(f"Sheet: {sheet_name}")
            for name, table in tables.items():
                print(f"  Table: {name!r}  ref: {table.ref}")
        else:
            print(f"Sheet: {sheet_name}  (no tables)")


def excel_to_text_grid_values_only(file_path, sheet_name=None):
    """Returns plain text content of a sheet — one line per row, values separated by spaces, no metadata."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Collect top-left coords of merged ranges so ghost cells are skipped
    ghost_cells = set()
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left = sheet.cell(row=min_row, column=min_col).coordinate
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                coord = sheet.cell(row=r, column=c).coordinate
                if coord != top_left:
                    ghost_cells.add(coord)

    lines = []
    for row in sheet.iter_rows():
        values = []
        for cell in row:
            if cell.coordinate in ghost_cells:
                continue
            if cell.value is None or str(cell.value).strip() == "":
                continue
            values.append(str(cell.value).strip())
        if values:
            lines.append("  ".join(values))

    return "\n".join(lines)


def excel_to_text_grid_full(file_path, sheet_name=None):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    # Check if the entire sheet is protected (prevents editing locked cells)
    is_sheet_protected = bool(sheet.protection and sheet.protection.sheet)

    # Pre-compute "ghost" merged cells
    ghost_cells = set()
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_coord = sheet.cell(row=min_row, column=min_col).coordinate
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                coord = sheet.cell(row=r, column=c).coordinate
                if coord != top_left_coord:
                    ghost_cells.add(coord)
    active_sheet_name = sheet.title
    checkboxes = extract_checkboxes(file_path, active_sheet_name)
    checkbox_map = {}
    for cb in checkboxes:
        key = (cb['row'], cb['col'])
        checkbox_map.setdefault(key, []).append(cb)

    def get_cell_color(c):
        try:
            fill = c.fill
            if fill and fill.start_color:
                color_obj = fill.start_color
                # Theme/indexed colors don't have a usable .rgb string
                if color_obj.type == 'rgb' and color_obj.rgb:
                    color = str(color_obj.rgb)
                    if color not in ['00000000', 'FFFFFFFF', 'None', '0']:
                        return color
        except (AttributeError, TypeError):
            pass
        return None

    # Find boundaries
    true_max_row = 1
    true_max_col = 1
    for row in sheet.iter_rows():
        for cell in row:
            has_value = cell.value is not None and str(
                cell.value).strip() != ""
            has_color = get_cell_color(cell) is not None
            has_border = cell.border and cell.border.bottom and cell.border.bottom.style
            is_explicitly_unlocked = cell.protection and not cell.protection.locked

            # If the cell has data, color, borders, or is explicitly unlocked, expand our grid!
            if has_value or has_color or has_border or is_explicitly_unlocked:
                if cell.row > true_max_row:
                    true_max_row = cell.row
                if cell.column > true_max_col:
                    true_max_col = cell.column

    true_max_row += 2
    true_max_col += 2

    grid_representation = ""

    for row in sheet.iter_rows(min_row=1, max_row=true_max_row, min_col=1, max_col=true_max_col):
        cell_data = []
        for cell in row:
            if cell.coordinate in ghost_cells:
                continue

            val = " ".join(str(cell.value).split()) if cell.value is not None else ""

            is_locked = bool(
                cell.protection and is_sheet_protected and cell.protection.locked)

            # 1. Color — [#RRGGBB], strip leading AA alpha byte
            color_meta = ""
            bg_color = get_cell_color(cell)
            if bg_color:
                hex_color = bg_color[2:] if len(bg_color) == 8 else bg_color
                color_meta = f" [#{hex_color}]"

            # 2. Protection — [L] for locked, nothing for unlocked
            prot_meta = ""
            if cell.protection:
                if is_sheet_protected and cell.protection.locked:
                    prot_meta = " [L]"

            # 3. Strikethrough — [S] means the text is struck through (not to be filled)
            strike_meta = ""
            try:
                if cell.font and cell.font.strike:
                    strike_meta = " [S]"
            except (AttributeError, TypeError):
                pass

            # 4. Checkboxes — [CHECKBOX ✓"label"] / [CHECKBOX ○"label"]
            checkbox_meta = ""
            for cb in checkbox_map.get((cell.row, cell.column), []):
                symbol = "✓" if cb['checked'] else "○"
                lbl = f'"{cb["label"]}"' if cb['label'] else ''
                checkbox_meta += f" [CHECKBOX {symbol}{lbl}]"

            cell_data.append(
                f"{cell.coordinate}:{val}{color_meta}{strike_meta}{checkbox_meta}{prot_meta}|")

        if cell_data:
            grid_representation += " ".join(cell_data) + "\n"

    return grid_representation
