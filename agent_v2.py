import pathlib
import re
import shutil
import sys
import openpyxl
from typing import Literal, Optional

from pydantic import BaseModel, Field
from dotenv import load_dotenv
from langchain.chat_models import init_chat_model
from langchain_core.messages import HumanMessage, SystemMessage
from langfuse.langchain import CallbackHandler

from excel_read_helpers import excel_to_text_grid_values_only, excel_to_text_grid_full
from excel_write_helpers import write_cells

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

load_dotenv()

llm = init_chat_model("google_genai:gemini-3.1-pro-preview", temperature=0)
langfuse_handler = CallbackHandler()


# ── Sheet categories ───────────────────────────────────────────────────────────

SheetCategory = Literal[
    "read_only",           # Scoring/evaluation sheets — skip entirely, evaluator-facing only
    "instructions",        # Participation conditions / fill instructions for the applicant
    "declaration",         # Legal declarations (Erklärung, Eigenerklärung Russland, etc.)
    "company_form",        # Main application form — company master data, staff, equipment, revenue
    "reference_company",   # One company/office project reference per sheet
    "reference_personnel", # One key staff member — qualifications + personal references
    "fee_offer",           # Fee/pricing offer — hourly rates, HOAI percentages, surcharges
]


# ── LLM classification schema ──────────────────────────────────────────────────

class SheetClassificationItem(BaseModel):
    sheet: str = Field(description="Exact sheet name as given")
    category: SheetCategory
    reference_slots: int = Field(
        default=1,
        description=(
            "Number of independent reference slots on this sheet. "
            "For reference_company: number of side-by-side reference columns "
            "('Referenz 1', 'Referenz 2', ...). Set to 1 for one-per-sheet layouts. "
            "For reference_personnel: number of project reference columns PER person. "
            "Set to 1 if each person has a single project reference block. "
            "For all other categories: always 1."
        )
    )
    person_slots: int = Field(
        default=1,
        description=(
            "Only relevant for reference_personnel sheets. "
            "Set to N > 1 when the sheet has multiple people side-by-side as columns "
            "(e.g. column headers 'Person 1', 'Person 2', 'Person 3'). "
            "Set to 1 for all other categories and for single-person-per-sheet layouts."
        )
    )
    reason: str = Field(description="One sentence explaining why")


class ClassifyResponse(BaseModel):
    sheets: list[SheetClassificationItem]


_CLASSIFY_SYSTEM_PROMPT = """You are classifying sheets from a German tender/procurement Excel workbook.

For each sheet assign exactly one of these categories:

- read_only
  Scoring matrices and evaluation sheets used only by the evaluator/jury.
  Examples: "Wertung Teilnahmewettbewerb", "Bewertungsschema Referenzen", "Wertung Zuschlagskriterien".
  The applicant fills nothing here.

- instructions
  Participation conditions and general fill instructions addressed TO the applicant.
  Examples: "Teilnahmebedingungen", "Hinweise-Anforderungen".
  These contain rules like which cells to fill, what attachments to include, deadlines, and form conventions.
  No applicant data is entered here, but the content is used as instructions for filling other sheets.

- declaration
  Legal declarations the applicant must sign, e.g.:
  "Erklärung zur freiberuflichen Tätigkeit", "Eigenerklärung Bezug Russland".
  Need only company name, address, and representative name.

- company_form
  Main application form asking for company master data:
  name, address, contact person, staff counts, software/equipment, insurance, revenue figures,
  project manager details.

- reference_company
  A form to submit one company/office project reference.
  Asks for project title, client, completion date, costs (KGR 300-400), criteria checkboxes.
  One sheet = one reference slot.

- reference_personnel
  A form about one specific key staff member:
  their qualifications, years of experience, and their personal project references.

- fee_offer
  A fee/pricing offer sheet asking for hourly rates, special service prices,
  HOAI phase percentages, renovation surcharge, overhead.

reference_slots
  Only relevant for reference_company and reference_personnel sheets.
  - For reference_company: count the number of side-by-side reference columns
    ("Referenz 1", "Referenz 2", "Referenz 3" ...). Set to 1 for one-per-sheet layouts.
  - For reference_personnel: count the number of project reference columns PER person.
    Set to 1 if each person has a single project reference block.
  - For every other category: set to 1.

person_slots
  Only relevant for reference_personnel sheets.
  Count how many distinct people columns appear side-by-side on the sheet
  (e.g. column headers "Person 1", "Person 2", "Person 3").
  Set to 1 for all other categories and for single-person-per-sheet layouts.

Return one entry per sheet in the order they were given."""


# ── Profile slicing schema ────────────────────────────────────────────────────

class ProfileSlice(BaseModel):
    category: SheetCategory
    content: Optional[str] = Field(
        description="Verbatim copy of the relevant section(s) from the company profile, "
                    "or null if no relevant data exists for this category."
    )


class ProfileSlicesResponse(BaseModel):
    slices: list[ProfileSlice]


_SLICE_SYSTEM_PROMPT = """You are extracting sections from a company profile to be used for filling German tender forms.

You will receive the full company profile and a list of sheet categories.
For each category, copy ALL company data that could be relevant for that sheet type VERBATIM — word for word, exactly as written.
Do NOT summarize, rephrase, shorten, or add anything.

- declaration
  Copy: company name, legal form, full address, VAT ID (Ust-IdNr.), commercial register entry,
  legal representative / contact person name, bidder type (Einzelbewerber/ARGE),
  Eignungsleihe (yes/no), and intended subcontracting (yes/no + description).
  Also copy verbatim any statements about exclusion criteria, criminal records, sanctions,
  or compliance declarations — these drive the Ja/Nein answers on declaration sheets.
  These are the fields declaration and self-declaration sheets ask for.

- reference_company
  Copy all company/project references verbatim, including every reference's title, client,
  location, project period, completion date, BGF, storeys, LPH performed, Honorarzone,
  Anlagengruppen, cost data by Kostengruppen, fees, and all special content flags
  (Schadstoffsanierung, Brandschutz, Strangsanierung, renewable energy, occupied building,
  public client).

- reference_personnel
  Copy all key personnel sections verbatim: for each person their name, title, employer,
  contact details, full qualification (degree, institution, year, additional qualifications),
  years of experience, core areas, and all personal project references with every field.

- read_only        Return null.
- instructions     Return null.
- company_form     Return null.
- fee_offer        Return null."""


def extract_profile_slices(company_profile: str) -> dict[SheetCategory, Optional[str]]:
    """
    Single LLM call that reads the full company profile and returns verbatim
    excerpts for each sheet category that needs profile data.
    """
    categories_needed = ["declaration", "reference_company", "reference_personnel"]
    categories_block = "\n".join(f"- {c}" for c in categories_needed)

    human_prompt = (
        f"Extract profile data for these categories:\n{categories_block}\n\n"
        f"[COMPANY PROFILE]\n{company_profile}"
    )

    print(f"  [slice]   extracting profile sections for {len(categories_needed)} categories...")

    structured_llm = llm.with_structured_output(ProfileSlicesResponse)
    response: ProfileSlicesResponse = structured_llm.invoke(
        [SystemMessage(content=_SLICE_SYSTEM_PROMPT),
         HumanMessage(content=human_prompt)],
        config={"callbacks": [langfuse_handler]},
    )

    result: dict[SheetCategory, Optional[str]] = {
        "read_only": None,
        "instructions": None,
        "fee_offer": None,
    }
    for item in response.slices:
        result[item.category] = item.content
        size = f"{len(item.content):,} chars" if item.content else "null"
        print(f"  [slice]   '{item.category}' → {size}")

    return result


# ── Instructions extraction ────────────────────────────────────────────────────

class WorkbookInstructions(BaseModel):
    cell_selection_rules: str = Field(
        description=(
            "Rules for identifying which cells must be filled. "
            "Include colour coding (e.g. yellow = mandatory, grey = optional), "
            "explicit fill markers (X, ja/nein, date formats), and any stated "
            "conventions about which fields are compulsory vs. optional."
        )
    )
    reference_rules: str = Field(
        description=(
            "Requirements the company references must satisfy. "
            "Include: how many references are required, the allowed time period "
            "(e.g. LPH 8 completed between date A and date B), and any minimum "
            "criteria each reference must meet (project type, size, public client, etc.)."
        )
    )
    personnel_requirements: str = Field(
        description=(
            "Which key persons must be named and what their minimum qualifications are. "
            "Include role names, required degree/title, minimum years of experience, "
            "and any other stated eligibility conditions."
        )
    )


_EXTRACT_INSTRUCTIONS_SYSTEM_PROMPT = """You are reading participation conditions and fill instructions from a German tender workbook.

Extract only the three fields below. Ignore everything else — submission deadlines, scoring processes,
evaluation methods, data protection notices, negotiation procedures, and contact information are NOT needed.

cell_selection_rules
  How to identify which cells must be filled.
  Look for: colour coding rules (yellow/grey cells), explicit statements about mandatory vs. optional fields,
  fill markers (X, ja/nein, specific date formats), attachment numbering conventions.

reference_rules
  Constraints the company project references must satisfy.
  Look for: how many references are required, the allowed completion time window,
  minimum criteria per reference (project type, Honorarzone, LPH completion, public client, size, etc.).

personnel_requirements
  Which key persons must be named and what qualifications they need.
  Look for: role names (Projektleiter, Bauleiter, etc.), required degree or title,
  minimum years of experience, specific certifications."""


def extract_workbook_instructions(
    file_path: str,
    classification: dict[str, SheetClassificationItem],
) -> Optional[WorkbookInstructions]:
    """
    Reads all 'instructions' sheets from the workbook and runs a single LLM call
    to extract only the three things that matter for filling: cell selection rules,
    reference requirements, and personnel requirements.
    Returns None if there are no instructions sheets.
    """
    instruction_sheets = [name for name, item in classification.items() if item.category == "instructions"]
    if not instruction_sheets:
        return None

    combined = ""
    for name in instruction_sheets:
        text = excel_to_text_grid_values_only(file_path, name)
        combined += f"\n=== {name} ===\n{text}\n"
        print(f"  [instructions] read '{name}' ({len(text):,} chars)")

    print(f"  [instructions] extracting rules from {len(instruction_sheets)} sheet(s)...")

    structured_llm = llm.with_structured_output(WorkbookInstructions)
    result: WorkbookInstructions = structured_llm.invoke(
        [SystemMessage(content=_EXTRACT_INSTRUCTIONS_SYSTEM_PROMPT),
         HumanMessage(content=combined.strip())],
        config={"callbacks": [langfuse_handler]},
    )

    print(f"  [instructions] cell_selection_rules : {len(result.cell_selection_rules):,} chars")
    print(f"  [instructions] reference_rules      : {len(result.reference_rules):,} chars")
    print(f"  [instructions] personnel_requirements: {len(result.personnel_requirements):,} chars")

    return result


# ── Assignment mapping ────────────────────────────────────────────────────────

class ReferenceAssignment(BaseModel):
    slot_id: str = Field(description="Exact slot identifier as given in [REFERENCE SLOTS TO FILL]")
    reference_name: Optional[str] = Field(
        description="Title or name of the company reference assigned to this slot, "
                    "exactly as it appears in the profile. Null if the slot should be left empty."
    )
    reason: str = Field(description="One sentence explaining why this reference fits this slot best")


class AssignReferencesResponse(BaseModel):
    assignments: list[ReferenceAssignment]


class PersonnelAssignment(BaseModel):
    slot_id: str = Field(description="Exact slot identifier as given in [PERSONNEL SLOTS TO FILL]")
    person_name: Optional[str] = Field(
        description="Full name of the team member assigned to this slot, "
                    "exactly as it appears in the profile. Null if the slot should be left empty."
    )
    reason: str = Field(description="One sentence explaining why this person fits this slot best")


class AssignPersonnelResponse(BaseModel):
    assignments: list[PersonnelAssignment]


_ASSIGN_REFERENCES_SYSTEM_PROMPT = """You are assigning company project references to reference slots in a German tender form.

Rules:
- Each slot gets exactly one reference. Each reference may only be used once.
- Pick the reference that best matches the slot given the tender's reference_rules.
- If there are more slots than references, assign null to the extra slots.
- Prefer references that meet the most criteria (project type, time period, public client, size, etc.).
- Slots named "ff", "fortfolgende", or similar are valid overflow slots for additional references —
  treat them as regular slots and assign a reference if one is available.

Return one assignment per slot using the exact slot identifier given."""


_ASSIGN_PERSONNEL_SYSTEM_PROMPT = """You are assigning key team members to personnel slots in a German tender form.

Rules:
- Each slot gets exactly one person.
- Slot IDs follow one of these patterns:
    "Sheet"                              — single person, single reference
    "Sheet / Referenz N"                 — single person, multiple project references (same person for all)
    "Sheet / Person N"                   — multiple people, one reference each (different person per Person N)
    "Sheet / Person N / Referenz M"      — multiple people, multiple references each
- Grouping rule: slots that share the same "Sheet / Person N" prefix belong to the SAME person.
  Example: "MySheet / Person 1 / Referenz 1" and "MySheet / Person 1 / Referenz 2" → same person.
- Different "Person N" indices on the same sheet → different people.
  Example: "MySheet / Person 1 / Referenz 1" and "MySheet / Person 2 / Referenz 1" → different people.
- For the legacy "Sheet / Referenz N" pattern (no Person prefix), all sub-slots belong to the SAME person.
  Example: "Planungsverantwortl. TA / Referenz 1", "Planungsverantwortl. TA / Referenz 2" → same person.
- When slots belong to DIFFERENT sheets, each person may only be used once across those sheets.
- Match the person whose role and expertise best fits what the slot is asking for.
  Read the slot description carefully — it usually states the required role (e.g. "Verantwortliche Planung TA",
  "Projektleiterin Objektplanung", "Stellvertretender Projektleiter").
- If there are more distinct-sheet/person slots than people, assign null to the extra slots.

Return one assignment per slot using the exact slot identifier given."""


def assign_references(
    slot_ids: list[str],
    references_profile: str,
    reference_rules: str,
) -> dict[str, Optional[str]]:
    """
    Single LLM call: assigns one unique company reference to each slot.
    slot_ids are strings like "Büroreferenz 1" or "Angaben Unternehmensreferenzen / Referenz 2".
    Returns {slot_id: reference_title_or_None}.
    """
    slots_block = "\n".join(f"- {s}" for s in slot_ids)
    human_prompt = (
        f"[REFERENCE SLOTS TO FILL]\n{slots_block}\n\n"
        f"[REFERENCE RULES FROM TENDER]\n{reference_rules}\n\n"
        f"[AVAILABLE COMPANY REFERENCES]\n{references_profile}"
    )

    print(f"  [assign]  assigning references to {len(slot_ids)} slot(s)...")
    structured_llm = llm.with_structured_output(AssignReferencesResponse)
    response: AssignReferencesResponse = structured_llm.invoke(
        [SystemMessage(content=_ASSIGN_REFERENCES_SYSTEM_PROMPT),
         HumanMessage(content=human_prompt)],
        config={"callbacks": [langfuse_handler]},
    )

    result: dict[str, Optional[str]] = {}
    for item in response.assignments:
        result[item.slot_id] = item.reference_name
        print(f"  [assign]  '{item.slot_id}' → {item.reference_name!r}  ({item.reason})")

    # Safety net for any slot the LLM missed
    for sid in slot_ids:
        if sid not in result:
            result[sid] = None
            print(f"  [assign]  '{sid}' → None (fallback)")

    return result


def assign_personnel(
    slot_ids: list[str],
    personnel_profile: str,
    personnel_requirements: str,
) -> dict[str, Optional[str]]:
    """
    Single LLM call: assigns one unique team member to each slot.
    Returns {slot_id: person_name_or_None}.
    """
    slots_block = "\n".join(f"- {s}" for s in slot_ids)
    human_prompt = (
        f"[PERSONNEL SLOTS TO FILL]\n{slots_block}\n\n"
        f"[PERSONNEL REQUIREMENTS FROM TENDER]\n{personnel_requirements}\n\n"
        f"[AVAILABLE TEAM MEMBERS]\n{personnel_profile}"
    )

    print(f"  [assign]  assigning personnel to {len(slot_ids)} slot(s)...")
    structured_llm = llm.with_structured_output(AssignPersonnelResponse)
    response: AssignPersonnelResponse = structured_llm.invoke(
        [SystemMessage(content=_ASSIGN_PERSONNEL_SYSTEM_PROMPT),
         HumanMessage(content=human_prompt)],
        config={"callbacks": [langfuse_handler]},
    )

    result: dict[str, Optional[str]] = {}
    for item in response.assignments:
        result[item.slot_id] = item.person_name
        print(f"  [assign]  '{item.slot_id}' → {item.person_name!r}  ({item.reason})")

    for sid in slot_ids:
        if sid not in result:
            result[sid] = None
            print(f"  [assign]  '{sid}' → None (fallback)")

    return result


# ── Cell schemas (shared across all fill steps) ───────────────────────────────

class FilledCell(BaseModel):
    cell: str = Field(description="Cell coordinate, e.g. 'C12'")
    value: Optional[str] = Field(
        description="Value to write, or null if data is unavailable. "
                    "Checkboxes use '1' to check, null to leave unchecked."
    )
    type: Literal["input", "checkbox"] = Field(
        description="'checkbox' for checkbox controls, 'input' for everything else"
    )


class FillResponse(BaseModel):
    cells: list[FilledCell]


# ── Checkbox type enforcement ──────────────────────────────────────────────────

def _apply_checkbox_fixes(cells: list[dict], file_path: str, sheet_name: str) -> list[dict]:
    """
    After _consensus_fill, force type='checkbox' on any cell the LLM returned at a
    known checkbox coordinate. This prevents it from landing in input_writes and
    corrupting the anchor cell via openpyxl. Checkboxes the LLM omitted or set to
    null are left untouched — their VML state is preserved by write_cells's
    unconditional snapshot/restore.
    """
    from excel_read_helpers import extract_checkboxes
    from openpyxl.utils.cell import get_column_letter

    cbs = extract_checkboxes(file_path, sheet_name)
    if not cbs:
        return cells

    checkbox_coords = {
        f"{get_column_letter(cb['col'])}{cb['row']}"
        for cb in cbs
    }

    print(f"  [checkbox] known coords: {sorted(checkbox_coords)}")

    fixed: list[dict] = []
    for cell in cells:
        if cell["cell"] in checkbox_coords:
            original_type = cell.get("type")
            fixed_cell = {**cell, "type": "checkbox"}
            fixed.append(fixed_cell)
            tag = "" if original_type == "checkbox" else f" (type was '{original_type}' → forced)"
            print(f"  [checkbox] {cell['cell']} value={cell.get('value')!r}{tag}")
        else:
            fixed.append(cell)

    checkbox_cells = [c for c in fixed if c.get("type") == "checkbox"]
    if not checkbox_cells:
        print(f"  [checkbox] no checkbox cells to write")

    return fixed


# ── Reference company fill ─────────────────────────────────────────────────────

def _content_str(response) -> str:
    """Normalise LLM response content to a plain string (Gemini returns a list)."""
    c = response.content
    if isinstance(c, list):
        return "".join(part if isinstance(part, str) else part.get("text", "") for part in c).strip()
    return c.strip()


_FILTER_REFERENCE_SYSTEM_PROMPT = """You are given a block of text containing multiple company project references.
Copy VERBATIM — word for word, exactly as written — only the section that matches the given reference name.
Do NOT summarize, rephrase, or add anything. Return only the matching section, nothing else."""


def filter_reference_data(full_references_slice: str, reference_name: str) -> str:
    """Extract just the one assigned reference verbatim from the full references slice."""
    print(f"  [filter]  extracting data for '{reference_name}'...")
    response = llm.invoke(
        [SystemMessage(content=_FILTER_REFERENCE_SYSTEM_PROMPT),
         HumanMessage(content=f"Reference to extract: {reference_name}\n\n{full_references_slice}")],
        config={"callbacks": [langfuse_handler]},
    )
    result = _content_str(response)
    print(f"  [filter]  got {len(result):,} chars")
    return result


_FILL_REFERENCE_SYSTEM_PROMPT = """You are filling a German tender reference sheet on behalf of a company.

You will receive:
- [CELL SELECTION RULES] — how to identify which cells must be filled (colour coding, conventions)
- [REFERENCE RULES] — criteria and constraints from the tender instructions
- [REFERENCE DATA] — the single project reference to use for this sheet
- [SHEET GRID] — the full cell data for this sheet

## Grid format
Each row: ROW_NUMBER: (COORD:value [tags...]), ...
- [#RRGGBB] background colour   [L] locked/read-only   [S] strikethrough — do not fill
- [CHECKBOX ✓"label"] checked   [CHECKBOX ○"label"] unchecked

## Your task
In one pass: identify every fillable cell AND assign its value from [REFERENCE DATA].

Which cells to fill:
- Apply [CELL SELECTION RULES] first
- Yellow ([#FFFF00] or similar) = mandatory input — always fill if data exists
- Skip locked [L] cells, strikethrough [S] cells, date auto-fields, page numbers
- Only include cells where an applicant would genuinely write data
- Stay inside the table boundary — if the nearest unlocked empty cell is far to the right
  of all label and content columns, it is outside the table; look for the correct input
  cell inside the table instead (usually a placeholder cell in the same row)

How to fill:
- Use only values from [REFERENCE DATA] — null if the data is not there
- Checkboxes [CHECKBOX]: "1" to check, null to leave unchecked
- ja/nein fields: write "ja" or "nein"
- Placeholder cells — an unlocked cell that already contains a format hint (such as
  "Datum:", "Beschreibung:", "ja -> Beschreibung:", "netto €") is an input field;
  write the actual value into it, overwriting the hint text
- All non-null values must be strings"""


def _consensus_fill(
    sheet_name: str,
    human_prompt: str,
    system_prompt: str,
    max_attempts: int = 10,
) -> list[dict]:
    """
    Repeat the fill LLM call until two attempts return the same (cell, type) set,
    then return that result. Falls back to cells seen in >50% of attempts if no
    consensus is reached within max_attempts.
    """
    from collections import Counter

    structured_llm = llm.with_structured_output(FillResponse)
    attempts: list[list[dict]] = []

    for attempt in range(1, max_attempts + 1):
        print(f"  [consensus] '{sheet_name}' attempt {attempt}/{max_attempts}")
        response: FillResponse = structured_llm.invoke(
            [SystemMessage(content=system_prompt),
             HumanMessage(content=human_prompt)],
            config={"callbacks": [langfuse_handler]},
        )
        cells = [
            {"sheet": sheet_name, "cell": c.cell, "value": c.value, "type": c.type}
            for c in response.cells
        ]

        current_fp = frozenset((c["cell"], c["type"]) for c in cells)
        for i, prev in enumerate(attempts):
            if frozenset((c["cell"], c["type"]) for c in prev) == current_fp:
                print(f"  [consensus] '{sheet_name}' — attempt {attempt} matches {i + 1}, {len(cells)} cells")
                return cells

        attempts.append(cells)

    # No consensus — keep cells that appeared in more than 50% of attempts
    print(f"  [consensus] '{sheet_name}' — no consensus after {max_attempts}, merging by frequency")
    threshold = max_attempts * 0.5
    cell_counts: Counter = Counter((c["cell"], c["type"]) for a in attempts for c in a)
    first_seen: dict[str, dict] = {}
    for a in attempts:
        for c in a:
            if c["cell"] not in first_seen:
                first_seen[c["cell"]] = c

    merged = [first_seen[coord] for (coord, _), count in cell_counts.items() if count > threshold]
    print(f"  [consensus] '{sheet_name}' — {len(merged)} cells kept (threshold >{threshold:.0f})")
    return merged


def fill_reference_company_sheet(
    file_path: str,
    out_path: str,
    sheet_name: str,
    reference_name: Optional[str],
    full_references_slice: str,
    cell_selection_rules: str,
    reference_rules: str,
    slot_label: Optional[str] = None,
) -> list[dict]:
    """
    Full fill pipeline for one reference_company sheet/slot:
      filter → consensus fill → write
    slot_label: when the sheet has multiple reference columns, the label of the target column
                (e.g. "Referenz 2"). The LLM is told to fill only that column.
    Returns the list of filled cells.
    """
    label = f"'{sheet_name}'" if not slot_label else f"'{sheet_name}' / {slot_label}"
    print(f"\n  -- {label} (reference_company) --")

    if not reference_name:
        print(f"  [skip]    no reference assigned, leaving blank")
        return []

    # 1. Filter: get only this reference's data from the full slice
    reference_data = filter_reference_data(full_references_slice, reference_name)

    # 2. Consensus fill: identify cells and assign values, loop until stable
    grid_text = excel_to_text_grid_full(file_path, sheet_name)
    slot_instruction = (
        f"[TARGET SLOT] Fill ONLY the column labeled '{slot_label}'. "
        f"Do not write into any other reference columns.\n\n"
        if slot_label else ""
    )
    human_prompt = (
        f"{slot_instruction}"
        f"[CELL SELECTION RULES]\n{cell_selection_rules}\n\n"
        f"[REFERENCE RULES]\n{reference_rules}\n\n"
        f"[REFERENCE DATA]\n{reference_data}\n\n"
        f"[SHEET GRID] (sheet: {sheet_name})\n{grid_text}"
    )

    consensus_key = f"{sheet_name}/{slot_label}" if slot_label else sheet_name
    cells = _consensus_fill(consensus_key, human_prompt, _FILL_REFERENCE_SYSTEM_PROMPT)
    cells = _apply_checkbox_fixes(cells, file_path, sheet_name)
    filled_count = sum(1 for c in cells if c["value"] is not None)
    print(f"  [fill]    {filled_count}/{len(cells)} cells filled")

    # 3. Write: update the output file
    write_cells(out_path, {sheet_name: cells})

    return cells


# ── Reference personnel fill ───────────────────────────────────────────────────

_FILTER_PERSONNEL_SYSTEM_PROMPT = """You are given a block of text containing multiple team member profiles.
Copy VERBATIM — word for word, exactly as written — only the section that matches the given person's name.
Do NOT summarize, rephrase, or add anything. Return only the matching section, nothing else."""


def filter_personnel_data(full_personnel_slice: str, person_name: str) -> str:
    """Extract just the one assigned person's profile verbatim from the full personnel slice."""
    print(f"  [filter]  extracting data for '{person_name}'...")
    response = llm.invoke(
        [SystemMessage(content=_FILTER_PERSONNEL_SYSTEM_PROMPT),
         HumanMessage(content=f"Person to extract: {person_name}\n\n{full_personnel_slice}")],
        config={"callbacks": [langfuse_handler]},
    )
    result = _content_str(response)
    print(f"  [filter]  got {len(result):,} chars")
    return result


_FILL_PERSONNEL_SYSTEM_PROMPT = """You are filling a German tender personnel sheet on behalf of a company.

You will receive:
- [CELL SELECTION RULES] — how to identify which cells must be filled (colour coding, conventions)
- [PERSONNEL REQUIREMENTS] — qualifications and experience criteria from the tender instructions
- [PERSONNEL DATA] — the single team member's profile to use for this sheet
- [SHEET GRID] — the full cell data for this sheet

## Grid format
Each row: ROW_NUMBER: (COORD:value [tags...]), ...
- [#RRGGBB] background colour   [L] locked/read-only   [S] strikethrough — do not fill
- [CHECKBOX ✓"label"] checked   [CHECKBOX ○"label"] unchecked

## Your task
In one pass: identify every fillable cell AND assign its value from [PERSONNEL DATA].

Which cells to fill:
- Apply [CELL SELECTION RULES] first
- Yellow ([#FFFF00] or similar) = mandatory input — always fill if data exists
- Skip locked [L] cells, strikethrough [S] cells, date auto-fields, page numbers
- Only include cells where an applicant would genuinely write data

How to fill:
- Use only values from [PERSONNEL DATA] — null if the data is not there
- Checkboxes [CHECKBOX]: "1" to check, null to leave unchecked
- ja/nein fields: write "ja" or "nein"
- All non-null values must be strings"""


def fill_reference_personnel_sheet(
    file_path: str,
    out_path: str,
    sheet_name: str,
    person_name: Optional[str],
    full_personnel_slice: str,
    cell_selection_rules: str,
    personnel_requirements: str,
    person_label: Optional[str] = None,
    ref_label: Optional[str] = None,
) -> list[dict]:
    """
    Full fill pipeline for one reference_personnel sheet/slot:
      filter → consensus fill → write
    person_label: when the sheet has multiple person columns, the label of the target column
                  (e.g. "Person 1"). None for single-person sheets.
    ref_label:    when the person has multiple project reference columns, the label of the target
                  column (e.g. "Referenz 2"). None when only one reference block exists.
    Returns the list of filled cells.
    """
    slot_label = " / ".join(filter(None, [person_label, ref_label])) or None
    label = f"'{sheet_name}'" if not slot_label else f"'{sheet_name}' / {slot_label}"
    print(f"\n  -- {label} (reference_personnel) --")

    if not person_name:
        print(f"  [skip]    no person assigned, leaving blank")
        return []

    # 1. Filter: get only this person's data from the full slice
    personnel_data = filter_personnel_data(full_personnel_slice, person_name)

    # 2. Consensus fill: identify cells and assign values, loop until stable
    grid_text = excel_to_text_grid_full(file_path, sheet_name)

    # Build the column-targeting instruction
    column_part = f"Fill ONLY the column labeled '{person_label}'. Do not write into any other person columns. " if person_label else ""

    # Build the reference-number instruction
    ref_num = None
    if ref_label:
        _m = re.search(r"\d+", ref_label)
        ref_num = int(_m.group()) if _m else None
    ref_part = f"Use this person's project reference #{ref_num} (the {ref_num}. reference listed in [PERSONNEL DATA]). Do NOT reuse a project reference that was assigned to another column — each column must contain a different project. " if ref_num else ""

    slot_instruction = (
        f"[TARGET SLOT] {column_part}{ref_part}\n\n"
        if (column_part or ref_part) else ""
    )

    human_prompt = (
        f"{slot_instruction}"
        f"[CELL SELECTION RULES]\n{cell_selection_rules}\n\n"
        f"[PERSONNEL REQUIREMENTS]\n{personnel_requirements}\n\n"
        f"[PERSONNEL DATA]\n{personnel_data}\n\n"
        f"[SHEET GRID] (sheet: {sheet_name})\n{grid_text}"
    )

    consensus_key = f"{sheet_name}/{slot_label}" if slot_label else sheet_name
    cells = _consensus_fill(consensus_key, human_prompt, _FILL_PERSONNEL_SYSTEM_PROMPT)
    cells = _apply_checkbox_fixes(cells, file_path, sheet_name)
    filled_count = sum(1 for c in cells if c["value"] is not None)
    print(f"  [fill]    {filled_count}/{len(cells)} cells filled")

    # 3. Write: update the output file
    write_cells(out_path, {sheet_name: cells})

    return cells


# ── Declaration fill ───────────────────────────────────────────────────────────

_FILL_DECLARATION_SYSTEM_PROMPT = """You are filling a German tender legal declaration sheet on behalf of a company.

You will receive:
- [CELL SELECTION RULES] — how to identify which cells must be filled (colour coding, conventions)
- [COMPANY DATA] — company name, legal form, address, and representative name
- [SHEET GRID] — the full cell data for this sheet

## Grid format
Each row: ROW_NUMBER: (COORD:value [tags...]), ...
- [#RRGGBB] background colour   [L] locked/read-only   [S] strikethrough — do not fill
- [CHECKBOX ✓"label"] checked   [CHECKBOX ○"label"] unchecked

## Your task
In one pass: identify every fillable cell AND assign its value from [COMPANY DATA].

Declaration sheets typically ask only for: company name, address, representative name/signature line,
and occasionally a ja/nein checkbox confirming a statement.

Which cells to fill:
- Apply [CELL SELECTION RULES] first
- Yellow ([#FFFF00] or similar) = mandatory input — always fill if data exists
- Skip locked [L] cells, strikethrough [S] cells, date/timestamp fields, page numbers

## Ja / Nein rows
When a row contains "Ja:" and "Nein:" as option labels, write "X" next to the correct label.
Never overwrite the label cell itself.
For questions about exclusion criteria, criminal records, sanctions, or disqualifying facts:
if [COMPANY DATA] contains no indication of any such issue, answer "Nein".

How to fill:
- Use only values from [COMPANY DATA] — null if the data is not there
- Checkboxes [CHECKBOX]: "1" to check, null to leave unchecked
- All non-null values must be strings"""


def fill_declaration_sheet(
    file_path: str,
    out_path: str,
    sheet_name: str,
    declaration_data: str,
    cell_selection_rules: str,
) -> list[dict]:
    """
    Fill pipeline for one declaration sheet:
      consensus fill (no filter needed — declaration data is already minimal) → write
    Returns the list of filled cells.
    """
    print(f"\n  -- '{sheet_name}' (declaration) --")

    grid_text = excel_to_text_grid_full(file_path, sheet_name)
    human_prompt = (
        f"[CELL SELECTION RULES]\n{cell_selection_rules}\n\n"
        f"[COMPANY DATA]\n{declaration_data}\n\n"
        f"[SHEET GRID] (sheet: {sheet_name})\n{grid_text}"
    )

    cells = _consensus_fill(sheet_name, human_prompt, _FILL_DECLARATION_SYSTEM_PROMPT)
    cells = _apply_checkbox_fixes(cells, file_path, sheet_name)
    filled_count = sum(1 for c in cells if c["value"] is not None)
    print(f"  [fill]    {filled_count}/{len(cells)} cells filled")

    write_cells(out_path, {sheet_name: cells})

    return cells


# ── Company form fill ─────────────────────────────────────────────────────────

_FILL_COMPANY_FORM_SYSTEM_PROMPT = """You are filling a German tender application form on behalf of a company.

You will receive:
- [CELL SELECTION RULES] — how to identify which cells must be filled (colour coding, conventions)
- [COMPANY DATA] — full company master data: name, address, contact, staff counts, financials, team
- [SHEET GRID] — the full cell data for this sheet

## Grid format
Each row: ROW_NUMBER: (COORD:value [tags...]), ...
- [#RRGGBB] background colour   [L] locked/read-only   [S] strikethrough — do not fill
- [CHECKBOX ✓"label"] checked   [CHECKBOX ○"label"] unchecked

## Your task
In one pass: identify every fillable cell AND assign its value from [COMPANY DATA].

These sheets typically ask for:
- Company name, legal form, address, phone, email, contact person
- Staff counts (architects, engineers, site supervision)
- Revenue figures per year and 3-year average — when the form asks for "Gesamtumsatz" (general/total
  turnover) use the general annual revenue figures, NOT the field-specific revenue figures.
  Only use field-specific revenue when the form explicitly asks for revenue in the relevant field.
- Insurance coverage amounts
- Project manager names and qualifications
- Subcontracting yes/no and description of scope
- ARGE yes/no

Which cells to fill:
- Apply [CELL SELECTION RULES] first
- Yellow ([#FFFF00] or similar) = mandatory input — always fill if data exists
- Skip locked [L] cells, strikethrough [S] cells, date/timestamp fields, page numbers
- If data for a cell is not present in [COMPANY DATA], return null — do NOT invent values

How to fill:
- Use only values explicitly present in [COMPANY DATA] — null if not there
- Checkboxes [CHECKBOX]: "1" to check, null to leave unchecked
- ja/nein fields: write "ja" or "nein"
- Multi-year columns: if the form asks for the same figure across several years (e.g. employee
  counts for 2022, 2023, 2024) and the profile only provides a total or average, use that
  figure as the best available value for each year column rather than leaving them blank
- Detail cells next to checkboxes: when a "ja" checkbox is accompanied by an unlocked text
  or description cell in the same row, fill that cell with the relevant supporting detail
  from [COMPANY DATA] (e.g. registration number, chamber membership, certification name)
- All non-null values must be strings"""


def fill_company_form_sheet(
    file_path: str,
    out_path: str,
    sheet_name: str,
    company_data: str,
    cell_selection_rules: str,
) -> list[dict]:
    """
    Fill pipeline for one company_form sheet:
      consensus fill (data already sliced) → write
    Returns the list of filled cells.
    """
    print(f"\n  -- '{sheet_name}' (company_form) --")

    grid_text = excel_to_text_grid_full(file_path, sheet_name)
    human_prompt = (
        f"[CELL SELECTION RULES]\n{cell_selection_rules}\n\n"
        f"[COMPANY DATA]\n{company_data}\n\n"
        f"[SHEET GRID] (sheet: {sheet_name})\n{grid_text}"
    )

    cells = _consensus_fill(sheet_name, human_prompt, _FILL_COMPANY_FORM_SYSTEM_PROMPT)
    cells = _apply_checkbox_fixes(cells, file_path, sheet_name)
    filled_count = sum(1 for c in cells if c["value"] is not None)
    print(f"  [fill]    {filled_count}/{len(cells)} cells filled")

    write_cells(out_path, {sheet_name: cells})

    return cells


# ── Fee offer fill ─────────────────────────────────────────────────────────────

_FILL_FEE_OFFER_SYSTEM_PROMPT = """You are filling a German tender fee offer sheet on behalf of a company.

You will receive:
- [CELL SELECTION RULES] — how to identify which cells must be filled (colour coding, conventions)
- [COMPANY DATA] — company master data (name, address, contact)
- [SHEET GRID] — the full cell data for this sheet

## Grid format
Each row: ROW_NUMBER: (COORD:value [tags...]), ...
- [#RRGGBB] background colour   [L] locked/read-only   [S] strikethrough — do not fill
- [CHECKBOX ✓"label"] checked   [CHECKBOX ○"label"] unchecked

## Your task
In one pass: identify every fillable cell AND assign its value from [COMPANY DATA].

Fee offer sheets ask for two types of data:
1. Company identity fields (name, address, representative) — fill from [COMPANY DATA]
2. Pricing fields (hourly rates, HOAI phase percentages, special service prices,
   renovation surcharge, overhead percentage) — leave null, these require human input

CRITICAL: Do NOT invent any pricing values. If a cell asks for a rate, percentage, or price
and the exact value is not present in [COMPANY DATA], return null.

Which cells to fill:
- Apply [CELL SELECTION RULES] first
- Yellow ([#FFFF00] or similar) = mandatory input — fill if data exists, null if not
- Skip locked [L] cells, strikethrough [S] cells, date/timestamp fields, page numbers

How to fill:
- Only use values explicitly present in [COMPANY DATA]
- All non-null values must be strings"""


def fill_fee_offer_sheet(
    file_path: str,
    out_path: str,
    sheet_name: str,
    company_data: str,
    cell_selection_rules: str,
) -> list[dict]:
    """
    Fill pipeline for one fee_offer sheet.
    Only fills company identity fields — pricing fields are left blank.
    Returns the list of filled cells.
    """
    print(f"\n  -- '{sheet_name}' (fee_offer) --")

    grid_text = excel_to_text_grid_full(file_path, sheet_name)
    human_prompt = (
        f"[CELL SELECTION RULES]\n{cell_selection_rules}\n\n"
        f"[COMPANY DATA]\n{company_data}\n\n"
        f"[SHEET GRID] (sheet: {sheet_name})\n{grid_text}"
    )

    cells = _consensus_fill(sheet_name, human_prompt, _FILL_FEE_OFFER_SYSTEM_PROMPT)
    cells = _apply_checkbox_fixes(cells, file_path, sheet_name)
    filled_count = sum(1 for c in cells if c["value"] is not None)
    print(f"  [fill]    {filled_count}/{len(cells)} cells filled (pricing fields left blank)")

    write_cells(out_path, {sheet_name: cells})

    return cells


# ── Classification ─────────────────────────────────────────────────────────────

def classify_sheets(file_path: str) -> dict[str, SheetClassificationItem]:
    """
    Read every visible sheet and classify it into a SheetCategory using a single LLM call.
    Sends only the first 30 lines of each sheet to keep the prompt small.
    Returns the full SheetClassificationItem (including reference_slots) keyed by sheet name.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    visible_sheets = [s.title for s in wb.worksheets if s.sheet_state == "visible"]
    wb.close()

    sheets_block = ""
    for name in visible_sheets:
        preview = excel_to_text_grid_values_only(file_path, name)
        preview_lines = "\n".join(preview.splitlines()[:30])
        sheets_block += f"\n=== SHEET: {name} ===\n{preview_lines}\n"

    print(f"  [classify] {len(visible_sheets)} sheets → asking LLM...")

    structured_llm = llm.with_structured_output(ClassifyResponse)
    response: ClassifyResponse = structured_llm.invoke(
        [SystemMessage(content=_CLASSIFY_SYSTEM_PROMPT),
         HumanMessage(content=sheets_block)],
        config={"callbacks": [langfuse_handler]},
    )

    result: dict[str, SheetClassificationItem] = {}
    for item in response.sheets:
        result[item.sheet] = item
        slots_note = f", {item.reference_slots} slots" if item.reference_slots > 1 else ""
        print(f"  [classify] '{item.sheet}' → {item.category}{slots_note}  ({item.reason})")

    # Safety net: any visible sheet not returned by LLM gets a safe default
    for name in visible_sheets:
        if name not in result:
            result[name] = SheetClassificationItem(sheet=name, category="company_form", reason="fallback")
            print(f"  [classify] '{name}' → company_form (fallback)")

    return result


# ── Entry point ────────────────────────────────────────────────────────────────

def run(file_path: str, only_sheet: Optional[str] = None) -> tuple[
    dict[str, SheetCategory],
    Optional[WorkbookInstructions],
    dict[SheetCategory, Optional[str]],
    dict[str, Optional[str]],
    dict[str, Optional[str]],
]:
    src = pathlib.Path(file_path)
    print(f"\n-- {src.name} --")

    company_profile = pathlib.Path("company_profile.md").read_text(encoding="utf-8")

    # Step 1: classify every sheet
    classification = classify_sheets(file_path)

    if only_sheet:
        classification = {k: v for k, v in classification.items() if k == only_sheet}
        print(f"\n  [filter] only_sheet='{only_sheet}'")

    print(f"\n  Classification summary:")
    for sheet, item in classification.items():
        slots_note = f", {item.reference_slots} slots" if item.reference_slots > 1 else ""
        print(f"    {sheet!r:45s} → {item.category}{slots_note}")

    # Step 2: extract structured instructions from the workbook's instructions sheets
    workbook_instructions = extract_workbook_instructions(file_path, classification)
    if workbook_instructions:
        print(f"\n  Workbook instructions extracted.")
    else:
        print(f"\n  No instructions sheets found.")

    # Step 3: extract the relevant profile slice for each category in use
    profile_slices = extract_profile_slices(company_profile)

    print(f"\n  Profile slices:")
    for cat, content in profile_slices.items():
        size = f"{len(content):,} chars" if content else "null"
        print(f"    {cat:25s} → {size}")

    # Step 4: build slot maps — expand multi-slot sheets into individual (sheet, slot_label) entries
    # ref_slots / personnel_slots: {slot_id -> (sheet_name, slot_label_or_None)}
    # slot_id is the string passed to the LLM for assignment and used as dict key.
    # slot_label is None for single-slot sheets (one-per-sheet layout).
    def _build_slot_map(category: SheetCategory) -> dict[str, tuple[str, Optional[str]]]:
        slot_map: dict[str, tuple[str, Optional[str]]] = {}
        for sheet_name, item in classification.items():
            if item.category != category:
                continue
            if category == "reference_personnel" and item.person_slots > 1:
                # Multi-person sheet: "Sheet / Person N / Referenz M" (or "Sheet / Person N" if refs=1)
                for p in range(item.person_slots):
                    person_label = f"Person {p + 1}"
                    if item.reference_slots > 1:
                        for r in range(item.reference_slots):
                            ref_label = f"Referenz {r + 1}"
                            slot_id = f"{sheet_name} / {person_label} / {ref_label}"
                            slot_map[slot_id] = (sheet_name, f"{person_label} / {ref_label}")
                    else:
                        slot_id = f"{sheet_name} / {person_label}"
                        slot_map[slot_id] = (sheet_name, person_label)
            elif item.reference_slots > 1:
                for i in range(item.reference_slots):
                    label = f"Referenz {i + 1}"
                    slot_id = f"{sheet_name} / {label}"
                    slot_map[slot_id] = (sheet_name, label)
            else:
                slot_map[sheet_name] = (sheet_name, None)
        return slot_map

    ref_slot_map = _build_slot_map("reference_company")
    personnel_slot_map = _build_slot_map("reference_personnel")

    # Step 5: assign references — one unique company reference per slot
    reference_assignment: dict[str, Optional[str]] = {}
    if ref_slot_map and profile_slices.get("reference_company"):
        reference_assignment = assign_references(
            slot_ids=list(ref_slot_map.keys()),
            references_profile=profile_slices["reference_company"],
            reference_rules=workbook_instructions.reference_rules if workbook_instructions else "",
        )

    # Step 6: assign personnel — one unique team member per slot
    personnel_assignment: dict[str, Optional[str]] = {}
    if personnel_slot_map and profile_slices.get("reference_personnel"):
        personnel_assignment = assign_personnel(
            slot_ids=list(personnel_slot_map.keys()),
            personnel_profile=profile_slices["reference_personnel"],
            personnel_requirements=workbook_instructions.personnel_requirements if workbook_instructions else "",
        )

    # Step 6: prepare output file
    out_path = str(pathlib.Path("tables_filled") / src.name)
    pathlib.Path("tables_filled").mkdir(exist_ok=True)
    shutil.copy2(file_path, out_path)

    cell_selection_rules = workbook_instructions.cell_selection_rules if workbook_instructions else ""
    reference_rules = workbook_instructions.reference_rules if workbook_instructions else ""
    personnel_requirements = workbook_instructions.personnel_requirements if workbook_instructions else ""

    all_filled: dict[str, list[dict]] = {}

    # Step 7: fill reference_company slots
    # Include declaration slice so the company name is available for the "Büro:" field
    ref_company_data = (
        (profile_slices.get("reference_company") or "")
        + "\n\n"
        + (profile_slices.get("declaration") or "")
    )
    for slot_id, (sheet_name, slot_label) in ref_slot_map.items():
        filled = fill_reference_company_sheet(
            file_path=file_path,
            out_path=out_path,
            sheet_name=sheet_name,
            reference_name=reference_assignment.get(slot_id),
            full_references_slice=ref_company_data,
            cell_selection_rules=cell_selection_rules,
            reference_rules=reference_rules,
            slot_label=slot_label,
        )
        all_filled[slot_id] = filled

    # Step 8: fill reference_personnel slots
    # Include declaration slice so the company name is available for the "Name des Bieters" field
    ref_personnel_data = (
        (profile_slices.get("reference_personnel") or "")
        + "\n\n"
        + (profile_slices.get("declaration") or "")
    )
    for slot_id, (sheet_name, slot_label) in personnel_slot_map.items():
        parts = slot_label.split(" / ") if slot_label else []
        person_label = next((p for p in parts if p.startswith("Person")), None)
        ref_label = next((p for p in parts if p.startswith("Referenz")), None)
        filled = fill_reference_personnel_sheet(
            file_path=file_path,
            out_path=out_path,
            sheet_name=sheet_name,
            person_name=personnel_assignment.get(slot_id),
            full_personnel_slice=ref_personnel_data,
            cell_selection_rules=cell_selection_rules,
            personnel_requirements=personnel_requirements,
            person_label=person_label,
            ref_label=ref_label,
        )
        all_filled[slot_id] = filled

    # Step 9: fill company_form sheets
    company_form_sheets = [s for s, item in classification.items() if item.category == "company_form"]
    for sheet_name in company_form_sheets:
        filled = fill_company_form_sheet(
            file_path=file_path,
            out_path=out_path,
            sheet_name=sheet_name,
            company_data=company_profile,
            cell_selection_rules=cell_selection_rules,
        )
        all_filled[sheet_name] = filled

    # Step 10: fill declaration sheets
    declaration_sheets = [s for s, item in classification.items() if item.category == "declaration"]
    for sheet_name in declaration_sheets:
        filled = fill_declaration_sheet(
            file_path=file_path,
            out_path=out_path,
            sheet_name=sheet_name,
            declaration_data=company_profile,
            cell_selection_rules=cell_selection_rules,
        )
        all_filled[sheet_name] = filled

    # Step 11: fill fee_offer sheets
    fee_sheets = [s for s, item in classification.items() if item.category == "fee_offer"]
    for sheet_name in fee_sheets:
        filled = fill_fee_offer_sheet(
            file_path=file_path,
            out_path=out_path,
            sheet_name=sheet_name,
            company_data=company_profile,
            cell_selection_rules=cell_selection_rules,
        )
        all_filled[sheet_name] = filled

    print(f"\n-- done: {out_path} --")
    return classification, workbook_instructions, profile_slices, reference_assignment, personnel_assignment


def run_test(
    file_path: str = r"tables/7d34b779-5fbd-47a3-ac59-04b69c9058ed.xlsx",
    sheet_name: str = "Eigenerklärung Bezug Russland",
) -> list[dict]:
    """
    Test helper: runs the full declaration fill pipeline for a single sheet only.
    Skips classification, slicing, assignment — goes straight to fill + write.
    Output is written to tables_filled/<original_filename>.
    """
    src = pathlib.Path(file_path)
    out_path = str(pathlib.Path("tables_filled") / src.name)
    pathlib.Path("tables_filled").mkdir(exist_ok=True)
    shutil.copy2(file_path, out_path)

    company_profile = pathlib.Path("company_profile.md").read_text(encoding="utf-8")
    workbook_instructions = extract_workbook_instructions(file_path, {})

    cell_selection_rules = workbook_instructions.cell_selection_rules if workbook_instructions else ""

    print(f"\n-- run_test: '{sheet_name}' in {src.name} --")
    cells = fill_declaration_sheet(
        file_path=file_path,
        out_path=out_path,
        sheet_name=sheet_name,
        declaration_data=company_profile,
        cell_selection_rules=cell_selection_rules,
    )
    print(f"\n-- done: {out_path} --")
    return cells


if __name__ == "__main__":
    # Usage:
    #   python agent_v2.py                                        — fill all files in tables/
    #   python agent_v2.py tables/foo.xlsx                        — fill one file
    #   python agent_v2.py tables/foo.xlsx "Sheet Name"           — fill one sheet only
    args = sys.argv[1:]
    if args and args[0].endswith(".xlsx"):
        only = args[1] if len(args) > 1 else None
        run(args[0], only_sheet=only)
    else:
        for f in sorted(pathlib.Path("tables").glob("*.xlsx")):
            run(str(f))

